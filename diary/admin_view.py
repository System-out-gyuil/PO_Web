from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db.models import Q, F, Sum
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import json
import os
import uuid
import boto3
import urllib.parse
from django.conf import settings
from botocore.exceptions import ClientError
from .models import Inquiry, Alarm, UserAlarm, User, Diary_main_count, Diary_diary_count, ClassForm, CountUser, CountUserIP
from django.core.paginator import Paginator
from django.core.serializers import serialize
from django.forms.models import model_to_dict
from diary.models import CalendarSettings, KanbanSettings, Attribute, Row, EmailVerification
import codecs

# 커스텀 JSON 인코더 클래스
class UnicodeJsonResponse(JsonResponse):
    def __init__(self, data, **kwargs):
        kwargs['json_dumps_params'] = {'ensure_ascii': False}
        super().__init__(data, **kwargs)

def save_uploaded_files(files):
    """업로드된 파일들을 S3에 저장하고 파일 정보를 반환"""
    saved_files = []
    
    # S3 클라이언트 생성
    s3_client = boto3.client(
        's3',
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        region_name=settings.AWS_S3_REGION_NAME
    )
    
    for uploaded_file in files:
        try:
            # 파일 확장자 추출
            file_ext = os.path.splitext(uploaded_file.name)[1]
            # 고유한 파일명 생성
            unique_filename = f"{uuid.uuid4()}{file_ext}"
            s3_key = f"alarm_files/{unique_filename}"
            
            # S3에 파일 업로드
            s3_client.upload_fileobj(
                uploaded_file,
                settings.AWS_STORAGE_BUCKET_NAME,
                s3_key,
                ExtraArgs={
                    'ContentType': uploaded_file.content_type,
                    'ContentDisposition': f'attachment; filename="{uploaded_file.name}"'
                }
            )
            
            # 다운로드 URL 생성
            download_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{s3_key}"
            
            # 서명된 다운로드 URL 생성 (24시간 유효)
            try:
                signed_download_url = s3_client.generate_presigned_url(
                    'get_object',
                    Params={
                        'Bucket': settings.AWS_STORAGE_BUCKET_NAME,
                        'Key': s3_key
                    },
                    ExpiresIn=86400  # 24시간
                )
            except Exception as e:
                print(f"서명된 URL 생성 실패: {e}")
                signed_download_url = download_url
            
            # 서명된 미리보기 URL 생성 (24시간 유효, inline으로 설정)
            try:
                signed_preview_url = s3_client.generate_presigned_url(
                    'get_object',
                    Params={
                        'Bucket': settings.AWS_STORAGE_BUCKET_NAME,
                        'Key': s3_key,
                        'ResponseContentDisposition': 'inline'
                    },
                    ExpiresIn=86400  # 24시간
                )
            except Exception as e:
                print(f"서명된 미리보기 URL 생성 실패: {e}")
                signed_preview_url = download_url
            
            # 파일 정보 저장
            file_info = {
                'original_name': uploaded_file.name,
                'saved_name': s3_key,
                'file_size': uploaded_file.size,
                'file_type': uploaded_file.content_type,
                'download_url': signed_download_url,
                'preview_url': signed_preview_url,
                'public_url': download_url
            }
            saved_files.append(file_info)
            
        except ClientError as e:
            error_code = e.response['Error']['Code']
            error_message = e.response['Error']['Message']
            print(f"S3 업로드 실패: {error_code} - {error_message}")
            # 실패한 파일은 건너뛰고 계속 진행
            continue
        except Exception as e:
            print(f"파일 업로드 중 오류: {e}")
            continue
    
    return saved_files

def admin_dashboard(request):
    """어드민 대시보드"""
    # 관리자 권한 확인
    user_id = request.session.get('diary_member_id')
    
    if not user_id:
        return redirect('login')
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return redirect('login')

    if not user.is_admin:  # is_admin 필드로 권한 확인
        return redirect('login')
    
    # 최근 문의사항 5개
    recent_inquiries = Inquiry.objects.all().order_by('-created_at')[:5]
    
    # 최근 공지사항 5개
    recent_alarms = Alarm.objects.all().order_by('-created_at')[:5]
    
    # 통계
    total_inquiries = Inquiry.objects.count()
    total_alarms = Alarm.objects.count()
    unread_inquiries = Inquiry.objects.filter(created_at__gte=timezone.now() - timezone.timedelta(days=7)).count()
    user_count = User.objects.count()
    
    # 조회수 통계 추가
    # 메인 페이지 총 조회수
    total_main_count = Diary_main_count.objects.aggregate(total=Sum('count'))['total'] or 0
    # 다이어리 페이지 총 조회수
    total_diary_count = Diary_diary_count.objects.aggregate(total=Sum('count'))['total'] or 0
    # 총 IP 수
    total_main_ips = Diary_main_count.objects.count()
    total_diary_ips = Diary_diary_count.objects.count()
    
    context = {
        'recent_inquiries': recent_inquiries,
        'recent_alarms': recent_alarms,
        'total_inquiries': total_inquiries,
        'total_alarms': total_alarms,
        'unread_inquiries': unread_inquiries,
        'user_count': user_count,
        'total_main_count': total_main_count,
        'total_diary_count': total_diary_count,
        'total_main_ips': total_main_ips,
        'total_diary_ips': total_diary_ips,
        'is_authenticated': True,  # 로그인 상태 추가
        'is_admin': user.is_admin,  # 관리자 상태 추가
        'user': user,  # 사용자 정보 추가
    }
    
    return render(request, 'diary/diary_admin.html', context)

def inquiry_list(request):
    """문의사항 목록 - JSON 응답"""
    user_id = request.session.get('diary_member_id')
    
    if not user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '사용자를 찾을 수 없습니다.'})

    if not user.is_admin:  # 실제로는 더 안전한 권한 체크 필요
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    # 검색 기능
    search_query = request.GET.get('search', '')
    inquiries = Inquiry.objects.all()
    
    if search_query:
        inquiries = inquiries.filter(
            Q(name__icontains=search_query) |
            Q(company_name__icontains=search_query) |
            Q(contact__icontains=search_query) |
            Q(content__icontains=search_query)
        )
    
    # 정렬
    sort_by = request.GET.get('sort', '-created_at')
    inquiries = inquiries.order_by(sort_by)
    
    # 페이지네이션
    paginator = Paginator(inquiries, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # JSON 응답을 위한 데이터 준비
    inquiries_data = []
    for inquiry in page_obj:
        inquiries_data.append({
            'id': inquiry.id,
            'name': inquiry.name,
            'company_name': inquiry.company_name,
            'contact': inquiry.contact,
            'content': inquiry.content,
            'created_at': inquiry.created_at.isoformat(),
        })
    
    pagination_data = {
        'number': page_obj.number,
        'num_pages': page_obj.paginator.num_pages,
        'has_previous': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
        'previous_page_number': page_obj.previous_page_number() if page_obj.has_previous() else None,
        'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None,
    }
    
    return JsonResponse({
        'success': True,
        'inquiries': inquiries_data,
        'pagination': pagination_data
    })

def inquiry_detail(request, inquiry_id):
    """문의사항 상세보기 - JSON 응답"""
    user_id = request.session.get('diary_member_id')
    
    if not user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '사용자를 찾을 수 없습니다.'})

    if not user.is_admin:  # 실제로는 더 안전한 권한 체크 필요
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    inquiry = get_object_or_404(Inquiry, id=inquiry_id)
    
    inquiry_data = {
        'id': inquiry.id,
        'name': inquiry.name,
        'company_name': inquiry.company_name,
        'contact': inquiry.contact,
        'content': inquiry.content,
        'created_at': inquiry.created_at.isoformat(),
    }
    
    return JsonResponse({
        'success': True,
        'inquiry': inquiry_data
    })

def alarm_list(request):
    """공지사항 목록 - JSON 응답"""
    user_id = request.session.get('diary_member_id')
    
    if not user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '사용자를 찾을 수 없습니다.'})

    if not user.is_admin:  # 실제로는 더 안전한 권한 체크 필요
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    # 검색 기능
    search_query = request.GET.get('search', '')
    alarms = Alarm.objects.all()
    
    if search_query:
        alarms = alarms.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query)
        )
    
    # 정렬
    sort_by = request.GET.get('sort', '-created_at')
    alarms = alarms.order_by(sort_by)
    
    # 페이지네이션
    paginator = Paginator(alarms, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # JSON 응답을 위한 데이터 준비
    alarms_data = []
    for alarm in page_obj:
        alarms_data.append({
            'id': alarm.id,
            'title': alarm.title,
            'content': alarm.get_text_content(),
            'files': alarm.get_files(),
            'created_at': alarm.created_at.isoformat(),
        })
    
    pagination_data = {
        'number': page_obj.number,
        'num_pages': page_obj.paginator.num_pages,
        'has_previous': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
        'previous_page_number': page_obj.previous_page_number() if page_obj.has_previous() else None,
        'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None,
    }
    
    return UnicodeJsonResponse({
        'success': True,
        'alarms': alarms_data,
        'pagination': pagination_data
    })

def alarm_create(request):
    """공지사항 작성"""
    user_id = request.session.get('diary_member_id')
    
    if not user_id:
        return redirect('login')
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return redirect('login')

    if not user.is_admin:  # 실제로는 더 안전한 권한 체크 필요
        return redirect('login')
    
    if request.method == 'POST':
        try:
            # JSON 데이터 처리
            if request.content_type == 'application/json':
                data = json.loads(request.body.decode('utf-8'))
                title = data.get('title')
                content_text = data.get('content')
                files_data = data.get('files', [])
            else:
                # Form 데이터 처리
                title = request.POST.get('title')
                content_text = request.POST.get('content')
                files_data = []
                
                # 파일 처리
                uploaded_files = request.FILES.getlist('files')
                if uploaded_files:
                    files_data = save_uploaded_files(uploaded_files)
        
        except json.JSONDecodeError:
            # JSON 파싱 실패 시 form data로 시도
            title = request.POST.get('title')
            content_text = request.POST.get('content')
            files_data = []
            
            # 파일 처리
            uploaded_files = request.FILES.getlist('files')
            if uploaded_files:
                files_data = save_uploaded_files(uploaded_files)
        
        if title and content_text is not None:
            # content를 dict 형태로 저장
            content = {
                'text': content_text,
                'files': files_data
            }
            
            alarm = Alarm.objects.create(
                title=title,
                content=content
            )
            
            # 모든 사용자에게 알람 생성
            users = User.objects.all()
            for user in users:
                UserAlarm.objects.create(
                    user=user,
                    alarm=alarm,
                    is_read=False
                )
            
            return UnicodeJsonResponse({'success': True, 'message': '공지사항이 성공적으로 작성되었습니다.'})
        else:
            return UnicodeJsonResponse({'success': False, 'message': '제목과 내용을 모두 입력해주세요.'})
    
    return render(request, 'diary/diary_admin.html', {
        'is_authenticated': True,
        'is_admin': user.is_admin,
        'user': user,
    })

def alarm_edit(request, alarm_id):
    """공지사항 수정"""
    user_id = request.session.get('diary_member_id')
    
    if not user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '사용자를 찾을 수 없습니다.'})

    if not user.is_admin:  # 실제로는 더 안전한 권한 체크 필요
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    alarm = get_object_or_404(Alarm, id=alarm_id)
    
    if request.method == 'POST':
        try:
            # JSON 데이터 처리
            if request.content_type == 'application/json':
                data = json.loads(request.body.decode('utf-8'))
                title = data.get('title')
                content_text = data.get('content')
                files_data = data.get('files', [])
            else:
                # Form 데이터 처리
                title = request.POST.get('title')
                content_text = request.POST.get('content')
                files_data = []
                
                # 파일 처리
                uploaded_files = request.FILES.getlist('files')
                if uploaded_files:
                    files_data = save_uploaded_files(uploaded_files)
        
        except json.JSONDecodeError:
            # JSON 파싱 실패 시 form data로 시도
            title = request.POST.get('title')
            content_text = request.POST.get('content')
            files_data = []
            
            # 파일 처리
            uploaded_files = request.FILES.getlist('files')
            if uploaded_files:
                files_data = save_uploaded_files(uploaded_files)
        
        if title and content_text is not None:
            # 기존 파일 정보 유지 (새 파일이 업로드되지 않은 경우)
            if not files_data:
                files_data = alarm.get_files()
            
            # content를 dict 형태로 저장
            content = {
                'text': content_text,
                'files': files_data
            }
            
            alarm.title = title
            alarm.content = content
            alarm.save()
            
            return UnicodeJsonResponse({'success': True, 'message': '공지사항이 성공적으로 수정되었습니다.'})
        else:
            return UnicodeJsonResponse({'success': False, 'message': '제목과 내용을 모두 입력해주세요.'})
    
    # GET 요청 시 알람 정보 반환
    alarm_data = {
        'id': alarm.id,
        'title': alarm.title,
        'content': alarm.get_text_content(),
        'files': alarm.get_files(),
        'created_at': alarm.created_at.isoformat(),
    }
    
    return UnicodeJsonResponse({
        'success': True,
        'alarm': alarm_data
    })

def alarm_delete(request, alarm_id):
    """공지사항 삭제"""
    user_id = request.session.get('diary_member_id')
    
    if not user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '사용자를 찾을 수 없습니다.'})

    if not user.is_admin:  # 실제로는 더 안전한 권한 체크 필요
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    alarm = get_object_or_404(Alarm, id=alarm_id)
    alarm.delete()
    
    return JsonResponse({'success': True, 'message': '공지사항이 삭제되었습니다.'})

@csrf_exempt
def inquiry_delete(request, inquiry_id):
    """문의사항 삭제"""
    user_id = request.session.get('diary_member_id')
    
    if not user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '사용자를 찾을 수 없습니다.'})

    if not user.is_admin:  # 실제로는 더 안전한 권한 체크 필요
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    inquiry = get_object_or_404(Inquiry, id=inquiry_id)
    inquiry.delete()
    
    return JsonResponse({'success': True, 'message': '문의사항이 삭제되었습니다.'})

@csrf_exempt
def admin_api(request):
    """어드민 API 엔드포인트"""
    user_id = request.session.get('diary_member_id')
    
    if not user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '사용자를 찾을 수 없습니다.'})

    if not user.is_admin:  # 실제로는 더 안전한 권한 체크 필요
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    if request.method == 'POST':
        data = json.loads(request.body)
        action = data.get('action')
        
        if action == 'create_alarm':
            title = data.get('title')
            content_text = data.get('content')
            files_data = data.get('files', [])
            
            if title and content_text:
                # content를 dict 형태로 저장
                content = {
                    'text': content_text,
                    'files': files_data
                }
                
                alarm = Alarm.objects.create(
                    title=title,
                    content=content
                )
                
                # 모든 사용자에게 알람 생성
                users = User.objects.all()
                for user in users:
                    UserAlarm.objects.create(
                        user=user,
                        alarm=alarm,
                        is_read=False
                    )
                
                return JsonResponse({
                    'success': True, 
                    'message': '공지사항이 성공적으로 작성되었습니다.',
                    'alarm_id': alarm.id
                })
            else:
                return JsonResponse({'success': False, 'message': '제목과 내용을 모두 입력해주세요.'})
        
        elif action == 'delete_alarm':
            alarm_id = data.get('alarm_id')
            try:
                alarm = Alarm.objects.get(id=alarm_id)
                alarm.delete()
                return JsonResponse({'success': True, 'message': '공지사항이 삭제되었습니다.'})
            except Alarm.DoesNotExist:
                return JsonResponse({'success': False, 'message': '존재하지 않는 공지사항입니다.'})
        
        elif action == 'delete_inquiry':
            inquiry_id = data.get('inquiry_id')
            try:
                inquiry = Inquiry.objects.get(id=inquiry_id)
                inquiry.delete()
                return JsonResponse({'success': True, 'message': '문의사항이 삭제되었습니다.'})
            except Inquiry.DoesNotExist:
                return JsonResponse({'success': False, 'message': '존재하지 않는 문의사항입니다.'})
    
    return JsonResponse({'success': False, 'message': '잘못된 요청입니다.'})

def user_list(request):
    """사용자 목록 API"""
    user_id = request.session.get('diary_member_id')
    
    if not user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '사용자를 찾을 수 없습니다.'})

    if not user.is_admin:
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    # 검색 및 정렬 파라미터
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort', '-created_at')
    page = request.GET.get('page', 1)
    
    # 사용자 쿼리셋
    users = User.objects.all()
    
    # 검색 필터링
    if search_query:
        users = users.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(company_name__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )
    
    # 정렬
    if sort_by:
        # Django ORM의 order_by는 - 접두사를 자동으로 처리합니다
        # use_date가 null인 경우를 고려하여 정렬
        if 'use_date' in sort_by:
            # use_date가 null인 경우를 마지막으로 정렬
            if sort_by.startswith('-'):
                users = users.order_by(F('use_date').desc(nulls_last=True))
            else:
                users = users.order_by(F('use_date').asc(nulls_last=True))
        else:
            users = users.order_by(sort_by)
    else:
        # 기본값: 최신순
        users = users.order_by('-created_at')
    
    # 페이지네이션
    paginator = Paginator(users, 20)  # 페이지당 20개
    try:
        users_page = paginator.page(page)
    except:
        users_page = paginator.page(1)
    
    # 사용자 데이터 직렬화
    users_data = []
    for user in users_page:
        user_dict = {
            'id': user.id,
            'name': user.name,
            'email': user.email,
            'company_name': user.company_name,
            'phone_number': user.phone_number,
            'created_at': user.created_at.isoformat(),
            'use_date': user.use_date.isoformat() if user.use_date else None,
            'is_admin': user.is_admin,
            'activate': user.activate  # 활성화 상태 추가
        }
        users_data.append(user_dict)
    
    # 페이지네이션 정보
    pagination = {
        'number': users_page.number,
        'num_pages': users_page.paginator.num_pages,
        'has_previous': users_page.has_previous(),
        'has_next': users_page.has_next(),
        'previous_page_number': users_page.previous_page_number() if users_page.has_previous() else None,
        'next_page_number': users_page.next_page_number() if users_page.has_next() else None,
    }
    
    return JsonResponse({
        'success': True,
        'users': users_data,
        'pagination': pagination,
        'current_user_id': user_id,  # 현재 로그인한 사용자 ID 추가
        'is_super_admin': user_id == 40  # 최고 관리자 여부 추가
    })

@csrf_exempt
def user_delete(request, user_id):
    """사용자 삭제"""
    admin_user_id = request.session.get('diary_member_id')
    
    if not admin_user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        admin_user = User.objects.get(id=admin_user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '관리자를 찾을 수 없습니다.'})

    if not admin_user.is_admin:
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    # 자기 자신을 삭제하려는 경우 방지
    if admin_user_id == user_id:
        return JsonResponse({'success': False, 'message': '자기 자신을 삭제할 수 없습니다.'})
    
    try:
        user_to_delete = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '삭제할 사용자를 찾을 수 없습니다.'})
    
    # 사용자와 관련된 모든 데이터 삭제
    try:
        # 1. UserAlarm 삭제
        UserAlarm.objects.filter(user=user_to_delete).delete()
        
        # 2. CalendarSettings 삭제
        CalendarSettings.objects.filter(user=user_to_delete).delete()
        
        # 3. KanbanSettings 삭제
        KanbanSettings.objects.filter(user=user_to_delete).delete()
        
        # 4. Attribute 삭제 (사용자별 속성)
        Attribute.objects.filter(user=user_to_delete).delete()
        
        # 5. Row 삭제 (사용자의 모든 행)
        Row.objects.filter(user=user_to_delete).delete()
        
        # 6. EmailVerification 삭제 (해당 사용자의 이메일 인증 데이터)
        EmailVerification.objects.filter(email=user_to_delete.email).delete()
        
        # 7. 사용자가 작성한 문의사항 삭제 (선택사항 - 필요에 따라 주석 처리)
        # Inquiry.objects.filter(user=user_to_delete).delete()
        
        # 8. 사용자 삭제
        user_to_delete.delete()
        
        return JsonResponse({'success': True, 'message': '사용자가 성공적으로 삭제되었습니다.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'사용자 삭제 중 오류가 발생했습니다: {str(e)}'})

@csrf_exempt
def user_toggle_admin(request, user_id):
    """사용자 관리자 권한 토글 - 최고 관리자(ID=1)만 가능"""
    admin_user_id = request.session.get('diary_member_id')
    
    if not admin_user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        admin_user = User.objects.get(id=admin_user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '관리자를 찾을 수 없습니다.'})

    # 최고 관리자(ID=1)만 권한 변경 가능
    if admin_user.id != 1:
        return JsonResponse({'success': False, 'message': '최고 관리자만 사용자 권한을 변경할 수 있습니다.'})
    
    # 자기 자신의 권한을 변경하려는 경우 방지
    if admin_user_id == user_id:
        return JsonResponse({'success': False, 'message': '자기 자신의 관리자 권한을 변경할 수 없습니다.'})
    
    try:
        user_to_toggle = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '변경할 사용자를 찾을 수 없습니다.'})
    
    try:
        # POST 데이터 파싱
        data = json.loads(request.body)
        make_admin = data.get('make_admin', False)
        
        # 관리자 권한 변경
        user_to_toggle.is_admin = make_admin
        user_to_toggle.save()
        
        status_text = '관리자' if make_admin else '일반 사용자'
        return JsonResponse({
            'success': True, 
            'message': f'사용자가 {status_text}로 변경되었습니다.',
            'new_status': 'admin' if make_admin else 'user'
        })
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': '잘못된 요청 형식입니다.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'관리자 권한 변경 중 오류가 발생했습니다: {str(e)}'})

@csrf_exempt
def user_toggle_activate(request, user_id):
    """사용자 계정 활성화/비활성화 토글"""
    admin_user_id = request.session.get('diary_member_id')
    
    if not admin_user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        admin_user = User.objects.get(id=admin_user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '관리자를 찾을 수 없습니다.'})

    if not admin_user.is_admin:
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    # 자기 자신의 활성화 상태를 변경하려는 경우 방지
    if admin_user_id == user_id:
        return JsonResponse({'success': False, 'message': '자기 자신의 계정 활성화 상태를 변경할 수 없습니다.'})
    
    try:
        user_to_toggle = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '변경할 사용자를 찾을 수 없습니다.'})
    
    try:
        # POST 데이터 파싱
        data = json.loads(request.body)
        make_active = data.get('make_active', False)
        
        # 활성화 상태 변경
        user_to_toggle.activate = make_active
        user_to_toggle.save()
        
        status_text = '활성화' if make_active else '비활성화'
        return JsonResponse({
            'success': True, 
            'message': f'사용자 계정이 {status_text}되었습니다.',
            'new_status': 'active' if make_active else 'inactive'
        })
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': '잘못된 요청 형식입니다.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'계정 활성화 상태 변경 중 오류가 발생했습니다: {str(e)}'})

@csrf_exempt
def user_update_use_date(request, user_id):
    """사용자 사용 기간 수정"""
    admin_user_id = request.session.get('diary_member_id')
    
    if not admin_user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        admin_user = User.objects.get(id=admin_user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '관리자를 찾을 수 없습니다.'})

    if not admin_user.is_admin:
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    try:
        user_to_update = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '수정할 사용자를 찾을 수 없습니다.'})
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            use_date_str = data.get('use_date')
            
            if use_date_str:
                from datetime import datetime
                use_date = datetime.strptime(use_date_str, '%Y-%m-%d').date()
                user_to_update.use_date = use_date
                user_to_update.save()
                
                return JsonResponse({
                    'success': True, 
                    'message': '사용 기간이 성공적으로 수정되었습니다.',
                    'use_date': use_date.isoformat()
                })
            else:
                return JsonResponse({'success': False, 'message': '사용 기간을 입력해주세요.'})
                
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': '잘못된 요청 형식입니다.'})
        except ValueError:
            return JsonResponse({'success': False, 'message': '올바른 날짜 형식이 아닙니다.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'사용 기간 수정 중 오류가 발생했습니다: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': '잘못된 요청입니다.'})

def diary_count_list(request):
    """다이어리 조회수 목록 API"""
    user_id = request.session.get('diary_member_id')
    
    if not user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '사용자를 찾을 수 없습니다.'})

    if not user.is_admin:
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    # 검색 및 정렬 파라미터
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort', '-count')
    page = request.GET.get('page', 1)
    count_type = request.GET.get('type', 'main')  # 'main' 또는 'diary'
    
    # 모델 선택
    if count_type == 'diary':
        count_model = Diary_diary_count
    else:
        count_model = Diary_main_count
    
    # 조회수 쿼리셋
    counts = count_model.objects.all()
    
    # 검색 필터링
    if search_query:
        counts = counts.filter(ip__icontains=search_query)
    
    # 정렬
    if sort_by:
        counts = counts.order_by(sort_by)
    else:
        # 기본값: 조회수 높은 순
        counts = counts.order_by('-count')
    
    # 페이지네이션
    total_ips = counts.count()  # 실제 IP 개수
    paginator = Paginator(counts, 20)  # 페이지당 20개
    try:
        counts_page = paginator.page(page)
    except:
        counts_page = paginator.page(1)
    
    # 조회수 데이터 직렬화
    counts_data = []
    for count in counts_page:
        count_dict = {
            'id': count.id,
            'ip': count.ip,
            'count': count.count,
            'created_at': count.created_at.isoformat(),
            'updated_at': count.updated_at.isoformat(),
        }
        counts_data.append(count_dict)
    
    # 페이지네이션 정보
    pagination = {
        'number': counts_page.number,
        'num_pages': counts_page.paginator.num_pages,
        'has_previous': counts_page.has_previous(),
        'has_next': counts_page.has_next(),
        'previous_page_number': counts_page.previous_page_number() if counts_page.has_previous() else None,
        'next_page_number': counts_page.next_page_number() if counts_page.has_next() else None,
    }
    
    # 총계 정보
    total_count = counts.aggregate(total=Sum('count'))['total'] or 0  # 조회수 합계
    
    return JsonResponse({
        'success': True,
        'counts': counts_data,
        'pagination': pagination,
        'total_count': total_count,
        'total_ips': total_ips,
        'count_type': count_type
    })

@csrf_exempt
def diary_count_delete(request, count_id):
    """다이어리 조회수 삭제"""
    admin_user_id = request.session.get('diary_member_id')
    
    if not admin_user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        admin_user = User.objects.get(id=admin_user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '관리자를 찾을 수 없습니다.'})

    if not admin_user.is_admin:
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    try:
        # count_type 파라미터로 어떤 모델에서 삭제할지 결정
        count_type = request.GET.get('type', 'main')
        if count_type == 'diary':
            count_to_delete = Diary_diary_count.objects.get(id=count_id)
        else:
            count_to_delete = Diary_main_count.objects.get(id=count_id)
        
        count_to_delete.delete()
        
        return JsonResponse({'success': True, 'message': '조회수 기록이 삭제되었습니다.'})
    except (Diary_main_count.DoesNotExist, Diary_diary_count.DoesNotExist):
        return JsonResponse({'success': False, 'message': '삭제할 조회수 기록을 찾을 수 없습니다.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'조회수 기록 삭제 중 오류가 발생했습니다: {str(e)}'})

def class_form_list(request):
    """클래스 신청 목록 API"""
    user_id = request.session.get('diary_member_id')
    
    if not user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '사용자를 찾을 수 없습니다.'})

    if not user.is_admin:
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    # 검색 및 정렬 파라미터
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort', '-created_at')
    page = request.GET.get('page', 1)
    
    # 클래스 신청 쿼리셋
    class_forms = ClassForm.objects.all()
    
    # 검색 필터링
    if search_query:
        class_forms = class_forms.filter(
            Q(name__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    # 정렬
    if sort_by:
        class_forms = class_forms.order_by(sort_by)
    else:
        # 기본값: 최신순
        class_forms = class_forms.order_by('-created_at')
    
    # 페이지네이션
    paginator = Paginator(class_forms, 20)  # 페이지당 20개
    try:
        class_forms_page = paginator.page(page)
    except:
        class_forms_page = paginator.page(1)
    
    # 클래스 신청 데이터 직렬화
    class_forms_data = []
    for class_form in class_forms_page:
        class_form_dict = {
            'id': class_form.id,
            'name': class_form.name,
            'phone': class_form.phone,
            'created_at': class_form.created_at.isoformat(),
            'updated_at': class_form.updated_at.isoformat(),
        }
        class_forms_data.append(class_form_dict)
    
    # 페이지네이션 정보
    pagination = {
        'number': class_forms_page.number,
        'num_pages': class_forms_page.paginator.num_pages,
        'has_previous': class_forms_page.has_previous(),
        'has_next': class_forms_page.has_next(),
        'previous_page_number': class_forms_page.previous_page_number() if class_forms_page.has_previous() else None,
        'next_page_number': class_forms_page.next_page_number() if class_forms_page.has_next() else None,
    }
    
    # 총계 정보
    total_count = class_forms.count()
    
    return JsonResponse({
        'success': True,
        'class_forms': class_forms_data,
        'pagination': pagination,
        'total_count': total_count
    })

@csrf_exempt
def class_form_delete(request, class_form_id):
    """클래스 신청 삭제"""
    admin_user_id = request.session.get('diary_member_id')
    
    if not admin_user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        admin_user = User.objects.get(id=admin_user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '관리자를 찾을 수 없습니다.'})

    if not admin_user.is_admin:
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    try:
        class_form_to_delete = ClassForm.objects.get(id=class_form_id)
        class_form_to_delete.delete()
        
        return JsonResponse({'success': True, 'message': '클래스 신청이 삭제되었습니다.'})
    except ClassForm.DoesNotExist:
        return JsonResponse({'success': False, 'message': '삭제할 클래스 신청을 찾을 수 없습니다.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'클래스 신청 삭제 중 오류가 발생했습니다: {str(e)}'})


def log_list(request):
    """로그 목록 API"""
    user_id = request.session.get('diary_member_id')
    
    if not user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '사용자를 찾을 수 없습니다.'})

    if not user.is_admin:
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    # 검색 및 정렬 파라미터
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort', '-created_at')
    page = request.GET.get('page', 1)
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # CountUserIP 쿼리셋
    logs = CountUserIP.objects.all()
    
    # 날짜 필터링
    if start_date and end_date:
        try:
            from datetime import datetime
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            # end_date는 해당 날짜의 마지막 시간까지 포함
            end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
            logs = logs.filter(created_at__range=[start_datetime, end_datetime])
        except ValueError:
            pass
    
    # 검색 필터링
    if search_query:
        logs = logs.filter(
            Q(ip__icontains=search_query) |
            Q(user__name__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )
    
    # 정렬
    if sort_by:
        logs = logs.order_by(sort_by)
    else:
        # 기본값: 최신순
        logs = logs.order_by('-created_at')
    
    # 페이지네이션
    paginator = Paginator(logs, 20)  # 페이지당 20개
    try:
        logs_page = paginator.page(page)
    except:
        logs_page = paginator.page(1)
    
    # 로그 데이터 직렬화
    logs_data = []
    for log in logs_page:
        log_dict = {
            'id': log.id,
            'ip': log.ip,
            'user_name': log.user.name if log.user else '알 수 없음',
            'user_email': log.user.email if log.user else '알 수 없음',
            'created_at': log.created_at.isoformat(),
            'updated_at': log.updated_at.isoformat(),
        }
        logs_data.append(log_dict)
    
    # 페이지네이션 정보
    pagination = {
        'number': logs_page.number,
        'num_pages': logs_page.paginator.num_pages,
        'has_previous': logs_page.has_previous(),
        'has_next': logs_page.has_next(),
        'previous_page_number': logs_page.previous_page_number() if logs_page.has_previous() else None,
        'next_page_number': logs_page.next_page_number() if logs_page.has_next() else None,
    }
    
    # 총계 정보
    total_count = logs.count()
    
    return JsonResponse({
        'success': True,
        'logs': logs_data,
        'pagination': pagination,
        'total_count': total_count
    })

def log_export_excel(request):
    """로그 데이터 엑셀 다운로드"""
    user_id = request.session.get('diary_member_id')
    
    if not user_id:
        return JsonResponse({'success': False, 'message': '로그인이 필요합니다.'})
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': '사용자를 찾을 수 없습니다.'})

    if not user.is_admin:
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        import io
        
        # 파라미터 가져오기
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        
        # CountUserIP 데이터 조회
        logs = CountUserIP.objects.all()
        
        # 날짜 필터링
        if start_date and end_date:
            try:
                from datetime import datetime
                start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                # end_date는 해당 날짜의 마지막 시간까지 포함
                end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
                logs = logs.filter(created_at__range=[start_datetime, end_datetime])
            except ValueError:
                pass
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "사용자 접속 로그"
        
        # 헤더 스타일 설정
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        headers = ["번호", "IP 주소", "사용자명", "이메일", "접속일"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 작성
        for row, log in enumerate(logs, 2):
            ws.cell(row=row, column=1, value=row-1)  # 번호
            ws.cell(row=row, column=2, value=log.ip)  # IP 주소
            ws.cell(row=row, column=3, value=log.user.name if log.user else '알 수 없음')  # 사용자명
            ws.cell(row=row, column=4, value=log.user.email if log.user else '알 수 없음')  # 이메일
            ws.cell(row=row, column=5, value=log.created_at.strftime('%Y-%m-%d %H:%M:%S'))  # 최초 접속
        
        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 파일명 생성
        if start_date and end_date:
            filename = f"사용자_접속_로그_{start_date}_to_{end_date}.xlsx"
        else:
            filename = f"사용자_접속_로그_전체_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # HTTP 응답 생성
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # 파일명을 UTF-8로 인코딩하여 헤더에 설정
        filename_encoded = urllib.parse.quote(filename)
        response['Content-Disposition'] = f'attachment; filename="{filename_encoded}"; filename*=UTF-8\'\'{filename_encoded}'
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'엑셀 생성 중 오류가 발생했습니다: {str(e)}'})
